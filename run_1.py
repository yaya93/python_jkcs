#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @FileName  :run_1.py
# @Time      :2020/7/3 22:38
# @Author    :yaya

from http_read1 import read_data
from http_read2 import http_request
from openpyxl  import load_workbook
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="gb18030")
Token=None

def run():
    global Token
    all_case=read_data('login.xlsx','recharge')
    print('获取到所测试数据是:',all_case)
    for i in range(len(all_case)):
        test_data=all_case[i]
        ip="http://120.78.128.25:8766"
        response=http_request(ip+test_data[4],eval(test_data[5]),token=Token,method=test_data[3])
        if 'login' in test_data[4]:
            Token='Bearer '+response['data']['token_info']['token']
        print("最后的结果值:",response)

        wb=load_workbook('login.xlsx')
        sheet=wb['recharge']
        sheet.cell(row=test_data[0]+1,column=8).value=str(response)
        #判断
        actual={'code':response['code'],'msg':response['msg']}
        if eval(test_data[6])==actual:
            print('测试用例通过')
            sheet.cell(row=test_data[0]+1,column=9).value='PASS'
        else:
            print('测试用例执行不通过')
            sheet.cell(row=test_data[0]+1,column=9).value='FAIL'

        #保存
        wb.save('login.xlsx')


run()



